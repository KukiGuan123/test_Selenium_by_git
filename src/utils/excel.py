import openpyxl
from pathlib import Path
from typing import List, Dict, Union


class ExcelUtil:
    """Excel 工具类：读取指定 sheet 的全部数据（行为字典）"""

    @staticmethod
    def get_sheet(excel_path: str, sheet_name: str) -> List[Dict[str, str | int | float | None]]:
        """
        获取整个 sheet 的所有数据（第一行作为 key，每一行作为字典）
        :param excel: Excel 文件路径
        :param sheet: Sheet 名称
        :return: [ {col1: val1, col2: val2}, {...} ]
        """
        if not Path(excel_path).exists():
            raise FileNotFoundError(f"Excel文件不存在: {excel_path}")

        # 加载工作簿
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet【{sheet_name}】不存在！可用sheet: {wb.sheetnames}")

        ws = wb[sheet_name]
        all_rows = list(ws.rows)
        if not all_rows:
            return []

        # 第一行作为表头 KEY
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in all_rows[0]]

        # 读取数据行
        result = []
        for row in all_rows[1:]:
            row_dict = {}
            for idx, cell in enumerate(row):
                row_dict[headers[idx]] = cell.value
            result.append(row_dict)

        wb.close()
        return result


    @staticmethod
    def get_row(
        path_Or_sheet: Union[str, List[Dict]],
        sheet_Or_filter: str = None,
        filter: str = None
    ) -> Dict:
        """
        两用方法：
        1. 传入 excel_path + sheet + filter
        2. 传入 sheet + filter
        """
        # 用法1：传入路径 (path, sheet, filter)
        if isinstance(path_Or_sheet, str) and sheet_Or_filter and filter:
            all_data = ExcelUtil.get_sheet(path_Or_sheet, sheet_Or_filter)
            for row in all_data:
                if str(row.get("Filter", "")).strip() == str(filter).strip():
                    return row
            return {}

        # 用法2：传入已经读取的数据 (sheet, filter)
        elif isinstance(path_Or_sheet, list) and sheet_Or_filter:
            all_data = path_Or_sheet
            for row in all_data:
                if str(row.get("Filter", "")).strip() == str(sheet_Or_filter).strip():
                    return row
            return {}

        else:
            raise ValueError("参数错误！用法：\n1. get_row(path, sheet, filter) \n2. get_row(data_list, filter)")


    @staticmethod
    def get_cell(
        path_Or_sheet_Or_row: Union[str, Dict, List[Dict]],
        sheet_Or_row_Or_filter: str = None,
        row_Or_filter: str = None,
        filter: str = None
    ):
        """
        🔥 支持 3 种用法（完全兼容旧代码）
        1. get_cell(path, sheet, filter, column)
        2. get_cell(row_dict, column)
        3. get_cell(sheet_data_list, filter, column)   👈 你要的新用法
        """

        # 用法1：4个参数 → path, sheet, filter, column
        if filter is not None:
            row = ExcelUtil.get_row(path_Or_sheet_Or_row, sheet_Or_row_Or_filter, row_Or_filter)
            return row.get(filter, None)

        # 用法2：2个参数 → row_dict, column
        if isinstance(path_Or_sheet_Or_row, dict) and sheet_Or_row_Or_filter is not None and row_Or_filter is None:
            return path_Or_sheet_Or_row.get(sheet_Or_row_Or_filter, None)

        # 用法3：3个参数 → sheet_data, filter, column
        if isinstance(path_Or_sheet_Or_row, list) and sheet_Or_row_Or_filter is not None and row_Or_filter is not None:
            row = ExcelUtil.get_row(path_Or_sheet_Or_row, sheet_Or_row_Or_filter)
            return row.get(row_Or_filter, None)

        # 都不满足
        raise ValueError("参数错误！支持用法：\n1. get_cell(path,sheet,filter,col)\n2. get_cell(row,col)\n3. get_cell(sheet_data,filter,col)")